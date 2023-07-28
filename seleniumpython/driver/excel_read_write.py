import openpyxl

# file="C:\\Users\\praka\\PycharmProjects\\seleniumpython\\python.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]

# READ DATA IN EXCEL SHEET
# row=sheet.max_row
# col=sheet.max_column
# print(row,"   ",col)
#
# for r in range(1,row+1):
#     for c in range(1,col+1):
#         print(sheet.cell(r,c).value,end="       ")
#     print()

# WRITE DATA IN EXCEL SHEET
file="C:\\Users\\praka\\PycharmProjects\\seleniumpython\\python.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]

# to write single data
# for r in range(1,4):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Welcome"
# workbook.save(file)

# to write multiple data
sheet.cell(1,1).value="name"
sheet.cell(1,2).value="address"

sheet.cell(2,1).value="vijay"
sheet.cell(2,2).value="CHN"

sheet.cell(3,1).value="kumar"
sheet.cell(3,2).value="BNG"

workbook.save(file)